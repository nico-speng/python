from lxml import etree
import urllib.request
import openpyxl


def createTable(data_list):
    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()

    # 创建一个新的工作表
    worksheet = workbook.active
    worksheet.title = '地址列表'

    # 添加表头
    worksheet.cell(row=1, column=1, value='地区')
    worksheet.cell(row=1, column=2, value='地址')
    
    for row_num, (key, value) in enumerate(data_list.items(), start=2):
        worksheet.cell(row=row_num, column=1, value=key)
        worksheet.cell(row=row_num, column=2, value=value)
    
    # 保存工作簿
    workbook.save('地址列表.xlsx')

url = 'https://www.meiguodizhi.com'

headers = {
    'accept-language':'zh-CN,zh;q=0.9',
    'content-type':'application/json',
    'cookie':'_gid=GA1.2.1652699743.1681279603; _ga=GA1.1.1743536245.1681279603; _ga_PE5DSSCM9B=GS1.1.1681284152.2.1.1681284632.0.0.0',
    'origin':'https://www.meiguodizhi.com',
    'referer':'https://www.meiguodizhi.com/kr-address',
    'sec-ch-ua':'"Chromium";v="112", "Google Chrome";v="112", "Not:A-Brand";v="99"',
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36',
}

request = urllib.request.Request(url=url,headers=headers)
response = urllib.request.urlopen(request)

# 获取网页源码
content = response.read().decode('utf-8')

# 解析网页源码
tree = etree.HTML(content)
# xpath 返回的都是list类型数据
list_name = tree.xpath('//div[@id="nav-top-1"]//a/text()')
list_str = tree.xpath('//div[@id="nav-top-1"]//a/@href')
base_url = []

for i in range(len(list_str)):
    strs =  url + list_str[i]
    base_url.append(strs)

# 合并两个列表为一个列表
combined_list = zip(list_name, base_url)
data_list = dict(combined_list)

# 创建表格
createTable(data_list)

import json
import requests

url = 'https://www.meiguodizhi.com/api/v1/dz'

payload = {
    'method':'address',
    'path':'/kr-address'
}

# 发送 POST 请求
response = requests.post(url,headers=headers,data=payload)

# 获取响应内容
content = response.content.decode()

# 尝试解析 JSON 响应
try:
    data = json.loads(content)
    print(data)
except json.JSONDecodeError:
    print("无法解析响应内容为 JSON 格式：", content)

